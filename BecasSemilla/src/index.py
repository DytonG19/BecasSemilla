import os
import io
import pandas as pd
from flask import Flask, render_template, request, flash, redirect, url_for, make_response
from config import config
from datetime import datetime
from openpyxl.styles import PatternFill, Font

app = Flask(__name__)
app.config['DEBUG'] = True
app.config['UPLOAD_FOLDER'] = 'uploads'

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

nombres_genericos = [
    'comercio', 'pastelería', 'repostería', 'panadería', 'mecánica', 'venta de verduras',
    'venta de ropa', 'venta de frutas', 'venta de productos', 'servicio de reparacion de motos',
    'floristería', 'agricultura', 'crianza de cerdos', 'crianza de pollos', 'crianza de gallinas ponedoras',
    'mini tienda', 'tienda', 'barbería', 'salón de belleza', 'electricidad', 'manualidades',
    'accesorios', 'ropa', 'muebles', 'fotografía', 'chef', 'diseñador de ropa', 'tortillería',
    'piñateria', 'piñatas', 'productos de limpieza', 'productos de bisuteria', 'productos desechabas',
    'productos desechables', 'refacciones', 'reparación de motocicletas', 'repuestos de luces para vehiculos',
    'auto car', 'lavado de autos', 'multiservicios', 'compra y venta de cerdos', 'venta de gelatina',
    'venta de papas fritas', 'venta de muebles', 'venta de frutas y jugo de naranja'
]

def calculate_age(birth_date_str):
    """Calcula la edad a partir de una fecha de nacimiento en string."""
    try:
        # Esta función de pandas es robusta para leer diferentes formatos numéricos (ej: dd/mm/yyyy, yyyy-mm-dd)
        birth_date = pd.to_datetime(birth_date_str, dayfirst=True, errors='coerce')
        if pd.isna(birth_date):
            return "N/A" # Si no se puede convertir, devuelve N/A
        today = datetime.today()
        age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
        return age
    except Exception:
        return "N/A"

def puntaje_marca(marca_val):
    if pd.isna(marca_val): return 5
    marca = str(marca_val).strip().lower()
    if (marca == '' or 'aun no' in marca or 'ningun' in marca or marca in nombres_genericos): return 5
    return 10

def calcular_puntajes_automaticos(df):
    puntajes = []
    COL_CURSO = 'Curso formacion en emprendimiento'
    COL_ESTADO = '2. Estado actual emprendimiento emprendimiento/negocio/proyecto:'
    COL_NOMBRE_MARCA = '2.1 Nombre o marca del emprendimiento'
    COL_REDES = '2.4 Página o redes sociales del emprendimiento:'
    COL_PRODUCTOS = '2.6 Mencione cuales son sus productos principales'
    for _, row in df.iterrows():
        puntos = {}
        curso_val = str(row.get(COL_CURSO, '')).strip().lower()
        puntos['curso'] = 0 if curso_val == 'no' or curso_val == '' else 10
        estado_val = str(row.get(COL_ESTADO, '')).lower()
        if 'funcionando' in estado_val: puntos['estado'] = 30
        elif 'iniciando' in estado_val: puntos['estado'] = 20
        elif 'idea definida' in estado_val: puntos['estado'] = 10
        else: puntos['estado'] = 0
        puntos['marca'] = puntaje_marca(row.get(COL_NOMBRE_MARCA, ''))
        redes_val = row.get(COL_REDES, '')
        puntos['redes'] = 5 if pd.isna(redes_val) or str(redes_val).strip() == '' or 'aun no' in str(redes_val).lower() else 10
        productos_val = row.get(COL_PRODUCTOS, '')
        puntos['productos'] = 10 if pd.isna(productos_val) or str(productos_val).strip() == '' or 'aun no' in str(productos_val).lower() else 20
        puntos['subtotal_auto'] = sum(puntos.values())
        puntajes.append(puntos)
    return pd.DataFrame(puntajes)

@app.route('/', methods=['GET', 'POST'])
def home():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No se encontró el campo de archivo.', 'error')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('No se seleccionó ningún archivo.', 'error')
            return redirect(request.url)
        if file and (file.filename.endswith('.xlsx')):
            try:
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
                file.save(filepath)
                df = pd.read_excel(filepath)
                df.columns = df.columns.str.strip()
                df_puntajes = calcular_puntajes_automaticos(df)
                df_completo = pd.concat([df, df_puntajes], axis=1)
                df_completo.fillna('', inplace=True)
                records = df_completo.to_dict('records')
                return render_template('resultados.html', aplicantes=records, show_form=True)
            except Exception as e:
                flash(f'Hubo un error al procesar el archivo: {e}', 'error')
                return redirect(request.url)
        else:
            flash('Formato de archivo no válido. Sube un archivo Excel (.xlsx)', 'error')
            return redirect(request.url)
    return render_template('index.html')

@app.route('/filtrar', methods=['POST'])
def filtrar():
    try:
        form_data = request.form.to_dict(flat=False)
        aplicantes_procesados = []
        num_aplicantes = len(form_data.get('id', []))
        for i in range(num_aplicantes):
            subtotal_auto = float(form_data.get('subtotal_auto', [])[i])
            puntaje_justificacion = float(form_data.get('puntaje_justificacion', [])[i] or 0)
            puntaje_total = subtotal_auto + puntaje_justificacion
            
            aplicante = {
                'ID': form_data.get('id', [])[i],
                'Nombre del joven': form_data.get('nombre', [])[i],
                'Puntaje Total': puntaje_total,
                'Puntaje Automático': subtotal_auto,
                'Puntaje Justificación': puntaje_justificacion,
                'Certificado Sinapsis': form_data.get('certificado_sinapsis', [])[i],
                'Fecha Nacimiento': form_data.get('fecha_nacimiento', [])[i],
                'Telefono': form_data.get('telefono', [])[i],
                'Email': form_data.get('email', [])[i],
                'Estado Emprendimiento': form_data.get('estado_emprendimiento', [])[i],
                'Puntaje Curso': form_data.get('puntaje_curso', [])[i],
                'Puntaje Estado': form_data.get('puntaje_estado', [])[i],
                'Puntaje Marca': form_data.get('puntaje_marca', [])[i],
                'Puntaje Redes': form_data.get('puntaje_redes', [])[i],
                'Puntaje Productos': form_data.get('puntaje_productos', [])[i],
                'Codigo CDI': form_data.get('codigo_cdi', [])[i],
                'Nombre CDI': form_data.get('nombre_cdi', [])[i],
                'Facilitador': form_data.get('facilitador', [])[i],
                'Padre Cuidador': form_data.get('padre_cuidador', [])[i],
                'Telefono Padre Cuidador': form_data.get('telefono_padre_cuidador', [])[i],
            }
            aplicantes_procesados.append(aplicante)
            
        aplicantes_ordenados = sorted(aplicantes_procesados, key=lambda x: x['Puntaje Total'], reverse=True)
        return render_template('resultados.html', aplicantes_ordenados=aplicantes_ordenados, show_results=True)
    except Exception as e:
        flash(f"Ocurrió un error inesperado al procesar los puntajes: {e}.", "error")
        return redirect(url_for('home'))

@app.route('/exportar-excel', methods=['POST'])
def exportar_excel():
    try:
        form_data = request.form.to_dict(flat=False)
        
        report_data = []
        num_aplicantes = len(form_data.get('id', []))
        for i in range(num_aplicantes):
            puntaje_total = float(form_data.get('puntaje_total', [])[i])
            report_data.append({
                'Ranking': int(float(form_data.get('ranking', [])[i])),
                'ID': form_data.get('id', [])[i],
                'Certificado Sinapsis': form_data.get('certificado_sinapsis', [])[i],
                'Nombre Completo': form_data.get('nombre', [])[i],
                'Edad': calculate_age(form_data.get('fecha_nacimiento', [])[i]),
                'Teléfono': form_data.get('telefono', [])[i],
                'Email': form_data.get('email', [])[i],
                'Código CDI': form_data.get('codigo_cdi', [])[i],
                'Nombre CDI': form_data.get('nombre_cdi', [])[i],
                'Facilitador': form_data.get('facilitador', [])[i],
                'Padre/Cuidador': form_data.get('padre_cuidador', [])[i],
                'Teléfono Padre/Cuidador': form_data.get('telefono_padre_cuidador', [])[i],
                'Estado Emprendimiento': form_data.get('estado_emprendimiento', [])[i],
                'Puntaje Curso': int(float(form_data.get('puntaje_curso', [])[i])),
                'Puntaje Estado': int(float(form_data.get('puntaje_estado', [])[i])),
                'Puntaje Marca': int(float(form_data.get('puntaje_marca', [])[i])),
                'Puntaje Redes': int(float(form_data.get('puntaje_redes', [])[i])),
                'Puntaje Productos': int(float(form_data.get('puntaje_productos', [])[i])),
                'Puntaje Justificación': int(float(form_data.get('puntaje_justificacion', [])[i])),
                'Puntaje Total': puntaje_total
            })
        
        df = pd.DataFrame(report_data)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Resultados')
            workbook = writer.book
            worksheet = writer.sheets['Resultados']

            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            blue_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)

            for cell in worksheet["1:1"]:
                cell.fill = header_fill
                cell.font = header_font

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=20, max_col=20):
                for cell in row:
                    if cell.value == 100:
                        cell.fill = green_fill
                    elif cell.value >= 90:
                        cell.fill = blue_fill
                    else:
                        cell.fill = yellow_fill
            
            column_widths = {'A': 10, 'B': 10, 'C': 25, 'D': 40, 'E': 8, 'F': 15, 'G': 30, 'H': 15, 'I': 25, 'J': 25, 'K': 30, 'L': 25, 'M': 30, 'N': 15, 'O': 15, 'P': 15, 'Q': 15, 'R': 15, 'S': 20, 'T': 15}
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

        output.seek(0)
        
        return make_response(output.getvalue(), 200, {
            'Content-Disposition': 'attachment; filename=Informe_Detallado_Resultados.xlsx',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })
        
    except Exception as e:
        flash(f"No se pudo exportar a Excel: {e}", "error")
        return redirect(url_for('home'))

if __name__ == '__main__':
    app.config.from_object(config['development'])
    app.run(debug=True,host="0.0.0.0", port=5000)
